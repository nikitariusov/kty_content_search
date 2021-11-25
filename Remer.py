import openpyxl  # для работы с эксель
import time
from Pars_sanremer import get_link, get_content
from Request import get_response, get_html

start_time = time.time()
excel_f = 'remer.xlsx'  # Передаем файл ексель
list_product = []  # список для словарей с информацией о товаре
HOST = 'http://sanremer.ru/'
HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                         'Chrome/90.0.4430.212 Safari/537.36', 'accept': '*/*'}


def read_excel(excel_file):  # Читаем нужные колонки в ексель
    global file_xl
    global excel_list

    file_xl = openpyxl.load_workbook(excel_file)  # важно
    excel_list = file_xl.active  # Активируем нужный лист в ексель

    print(f'Количество товара: {excel_list.max_row - 1}')  # показываем общее кол-во товара (строк в файле)

    for i in range(1, excel_list.max_row):  # Цикл по каждой строке таблицы
        dict_product = {}
        for cols_id in excel_list.iter_cols(1, 1):  # Цикл по каждой ячейке в колонке 1
            dict_product['id'] = cols_id[i].value  # добаВляем содержимое ячейки в словарь
            for cols_art in excel_list.iter_cols(12, 12):  # Цикл по каждой ячейке в колонке 12
                dict_product['art'] = cols_art[i].value  # добаВляем содержимое ячейки в словарь
                for cols_name in excel_list.iter_cols(3, 3):  # Цикл по каждой ячейке в колонке 3
                    dict_product['name'] = cols_name[i].value  # добаВляем содержимое ячейки в словарь

        list_product.append(dict_product)  # добаВляем словарь в список с товарами

    return list_product


def script_execution_time():
    runtime = time.time() - start_time
    print(f"[INFO] Время выполения: {'{0:.2f}'.format(runtime)}c.")


def create_content(description, name):
    end = f'<p>Купить {name[0].lower() + name[1:]}в оригинальной сборке и с официальной гарантией от производителя ' \
          f'можно онлайн в нашем  интернет-магазине и получить товар в любом городе Украины. ' \
          'Также возможен самовывоз.</p>'
    content = f'<p>{description}</p> {end}'
    return content


def recording_on_file(produkts, file):
    file_xl = openpyxl.load_workbook(file)  # важно
    excel_list = file_xl.active  # Активируем нужный лист в ексель

    row = 2
    for product in produkts:
        excel_list[f'W{row}'] = product.get('content')
        row += 1

    file_xl.save(f'Обработан_{file}')
    print(f'\nЗапись завершена. \nФайл сохранен: "Обработан_{file}"')


produkts = read_excel(excel_f)
# print(produkts)

n = 0
for product in produkts:
    #   print('[TEST]', product)
    art = product.get('art')
    #   print('[TEST]', art)
    print(f'[INFO] Поиск товара {n}...')

    url = HOST + 'search?&search_query=' + art  # собираем ссылку по которой будет запрос
    #   print('[TEST]', url)

    request = get_html(url, HEADERS)
    #   print('[TEST]', request.text)

    response = get_response(request)
    #   print('[TEST]', response)

    link = get_link(response)
    product['link'] = link
    n += 1
 
    if link:
        description = get_content(link, HEADERS)
        product_name = product.get('name')
        content = create_content(description, product_name)
        if content:
            print(f'\t[STATUS] OK')
        #   print('[TEST]', content)
        product['content'] = content
    else:
        print(f'\t[STATUS] FAILED: no link.')
recording_on_file(produkts, excel_f)

#   print('[TEST]', produkts)
script_execution_time()
